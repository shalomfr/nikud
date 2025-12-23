"""
מודול ייצוא לאקסל
Excel Exporter Module
"""

import io
from typing import List, Dict, Any
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelExporter:
    """מחלקה לייצוא נתונים לאקסל"""

    def __init__(self):
        # Define styles
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.rtl_alignment = Alignment(horizontal="right", vertical="center", reading_order=2)
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def prepare_dataframe(self, results: List[Dict]) -> pd.DataFrame:
        """הכנת DataFrame מהתוצאות"""
        data = []
        for result in results:
            row = {
                'מילה': result.get('word', ''),
                'מילה ללא ניקוד': result.get('word_plain', ''),
                'סוג הברה': result.get('syllable_type', ''),
                'תבנית ניקוד': result.get('nikud_pattern', ''),
                'יש שווא': 'כן' if result.get('has_shva') else 'לא',
                'סוגי שווא': ', '.join(result.get('shva_types', [])),
                'סימני ניקוד': ', '.join(result.get('nikud_marks', [])),
                'יש דגש': 'כן' if result.get('has_dagesh') else 'לא',
                'הברה פתוחה': 'כן' if result.get('has_open_syllable') else 'לא',
                'הברה סגורה': 'כן' if result.get('has_closed_syllable') else 'לא',
                'מקרים מיוחדים': ', '.join(result.get('special_cases', [])),
                'הקשר': result.get('context', ''),
                'מקור': result.get('source_name', ''),
                'קטגוריה': result.get('category_name', '')
            }
            data.append(row)

        return pd.DataFrame(data)

    def format_worksheet(self, worksheet, df: pd.DataFrame, add_filters: bool = True):
        """עיצוב גיליון אקסל"""
        # Set column widths
        column_widths = {
            'A': 15,  # מילה
            'B': 15,  # מילה ללא ניקוד
            'C': 10,  # סוג הברה
            'D': 20,  # תבנית ניקוד
            'E': 8,   # יש שווא
            'F': 15,  # סוגי שווא
            'G': 20,  # סימני ניקוד
            'H': 8,   # יש דגש
            'I': 10,  # הברה פתוחה
            'J': 10,  # הברה סגורה
            'K': 15,  # מקרים מיוחדים
            'L': 30,  # הקשר
            'M': 15,  # מקור
            'N': 15,  # קטגוריה
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Format header row
        for cell in worksheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

        # Format data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = self.rtl_alignment
                cell.border = self.thin_border

        # Add auto-filter
        if add_filters and worksheet.max_row > 1:
            worksheet.auto_filter.ref = f"A1:N{worksheet.max_row}"

        # Set RTL direction
        worksheet.sheet_view.rightToLeft = True

    def add_statistics_sheet(self, workbook: Workbook, df: pd.DataFrame):
        """הוספת גיליון סטטיסטיקות"""
        ws = workbook.create_sheet("סטטיסטיקות")
        ws.sheet_view.rightToLeft = True

        stats = [
            ("סטטיסטיקות כלליות", ""),
            ("", ""),
            ("סך הכל מילים", len(df)),
            ("מילים ייחודיות", df['מילה'].nunique()),
            ("מילים עם שווא", len(df[df['יש שווא'] == 'כן'])),
            ("מילים עם דגש", len(df[df['יש דגש'] == 'כן'])),
            ("", ""),
            ("התפלגות סוגי הברות", ""),
        ]

        # Add syllable distribution
        syllable_counts = df['סוג הברה'].value_counts()
        for syllable_type, count in syllable_counts.items():
            stats.append((syllable_type, count))

        # Write stats
        for row_idx, (label, value) in enumerate(stats, 1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)

        # Format
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

        # Style header
        ws['A1'].font = Font(bold=True, size=14)

    def add_info_sheet(self, workbook: Workbook):
        """הוספת גיליון מידע"""
        ws = workbook.create_sheet("מידע")
        ws.sheet_view.rightToLeft = True

        info = [
            ("מערכת ניתוח ניקוד - מידע על הקובץ", ""),
            ("", ""),
            ("תאריך יצירה", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("הסבר על העמודות:", ""),
            ("מילה", "המילה עם הניקוד המלא"),
            ("מילה ללא ניקוד", "המילה ללא סימני ניקוד"),
            ("סוג הברה", "פתוחה / סגורה / לא ידוע"),
            ("תבנית ניקוד", "תבנית סימני הניקוד במילה"),
            ("יש שווא", "האם יש שווא במילה"),
            ("סוגי שווא", "נע / נח / שני שווא וכו'"),
            ("סימני ניקוד", "רשימת סימני הניקוד במילה"),
            ("יש דגש", "האם יש דגש במילה"),
            ("מקרים מיוחדים", "קמץ קטן, פתח גנובה וכו'"),
            ("הקשר", "המשפט שבו מופיעה המילה"),
            ("מקור", "הטקסט המקורי"),
            ("קטגוריה", "קטגוריית הטקסט"),
        ]

        for row_idx, (label, value) in enumerate(info, 1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

        ws['A1'].font = Font(bold=True, size=14)

    def export_to_bytes(self, results: List[Dict]) -> bytes:
        """
        ייצוא התוצאות לאקסל והחזרת bytes
        Export results to Excel and return bytes
        """
        df = self.prepare_dataframe(results)

        # Create workbook
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write main sheet
            df.to_excel(writer, sheet_name='תוצאות חיפוש', index=False)

            # Get workbook and format
            workbook = writer.book
            worksheet = writer.sheets['תוצאות חיפוש']
            self.format_worksheet(worksheet, df)

            # Add additional sheets
            self.add_statistics_sheet(workbook, df)
            self.add_info_sheet(workbook)

        output.seek(0)
        return output.getvalue()


# Singleton instance
excel_exporter = ExcelExporter()

