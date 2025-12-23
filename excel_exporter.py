"""
מודול ייצוא תוצאות לאקסל
Excel Export Module for Search Results
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional
from datetime import datetime
import json

class ExcelExporter:
    """מייצא תוצאות חיפוש לקובץ אקסל מעוצב"""

    def __init__(self):
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.rtl_alignment = Alignment(horizontal="right", vertical="center")
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export_to_excel(self, results: List[Dict], output_path: str,
                       add_filters: bool = True, add_pivot: bool = False):
        """ייצוא תוצאות חיפוש לאקסל"""

        # יצירת DataFrame
        df = self.prepare_dataframe(results)

        # יצירת קובץ אקסל
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # גיליון עם התוצאות
            df.to_excel(writer, sheet_name='תוצאות חיפוש', index=False)

            # קבלת הגיליון לעיצוב
            workbook = writer.book
            worksheet = writer.sheets['תוצאות חיפוש']

            # עיצוב הגיליון
            self.format_worksheet(worksheet, df, add_filters)

            # הוספת גיליון סטטיסטיקות
            self.add_statistics_sheet(workbook, df)

            # הוספת טבלת Pivot אם נדרש
            if add_pivot:
                self.add_pivot_sheet(workbook, df)

            # הוספת גיליון עם הסברים
            self.add_info_sheet(workbook)

        print(f"הקובץ נשמר בהצלחה: {output_path}")

    def prepare_dataframe(self, results: List[Dict]) -> pd.DataFrame:
        """הכנת DataFrame מהתוצאות"""
        data = []

        for r in results:
            row = {
                'מילה': r.get('word', ''),
                'מילה ללא ניקוד': r.get('word_plain', ''),
                'סוג הברה': r.get('syllable_type', ''),
                'תבנית ניקוד': r.get('nikud_pattern', ''),
                'יש שווא': 'כן' if r.get('has_shva') else 'לא',
                'סוגי שווא': ', '.join(r.get('shva_types', [])) if r.get('shva_types') else '',
                'סימני ניקוד': self.format_nikud_marks(r.get('nikud_marks', [])),
                'יש דגש': 'כן' if r.get('has_dagesh') else 'לא',
                'הברה פתוחה': 'כן' if r.get('has_open_syllable') else 'לא',
                'הברה סגורה': 'כן' if r.get('has_closed_syllable') else 'לא',
                'מקרים מיוחדים': ', '.join(r.get('special_cases', [])),
                'הקשר': r.get('context', ''),
                'מקור': r.get('source_name', ''),
                'קטגוריה': r.get('category_name', ''),
                'מיקום': r.get('position', 0)
            }
            data.append(row)

        return pd.DataFrame(data)

    def format_nikud_marks(self, marks: List[str]) -> str:
        """עיצוב רשימת סימני ניקוד"""
        nikud_names = {
            '\u05B0': 'שווא',
            '\u05B1': 'חטף סגול',
            '\u05B2': 'חטף פתח',
            '\u05B3': 'חטף קמץ',
            '\u05B4': 'חיריק',
            '\u05B5': 'צירה',
            '\u05B6': 'סגול',
            '\u05B7': 'פתח',
            '\u05B8': 'קמץ',
            '\u05B9': 'חולם',
            '\u05BA': 'חולם מלא',
            '\u05BB': 'קובוץ',
            '\u05BC': 'דגש/שורוק',
            '\u05BD': 'מתג',
            '\u05BF': 'רפה',
            '\u05C1': 'שין ימנית',
            '\u05C2': 'שין שמאלית'
        }

        named_marks = []
        for mark in marks:
            if mark in nikud_names:
                named_marks.append(nikud_names[mark])
            else:
                named_marks.append(mark)

        return ', '.join(named_marks)

    def format_worksheet(self, worksheet, df: pd.DataFrame, add_filters: bool):
        """עיצוב גיליון העבודה"""
        # הגדרת RTL לגיליון
        worksheet.sheet_view.rightToLeft = True

        # עיצוב הכותרות
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border

        # עיצוב התאים
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = self.rtl_alignment
                cell.border = self.border

        # התאמת רוחב עמודות
        for col_idx, col_name in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_idx)
            if col_name == 'הקשר':
                worksheet.column_dimensions[column_letter].width = 50
            elif col_name in ['מילה', 'מילה ללא ניקוד']:
                worksheet.column_dimensions[column_letter].width = 15
            elif col_name == 'מקרים מיוחדים':
                worksheet.column_dimensions[column_letter].width = 25
            else:
                worksheet.column_dimensions[column_letter].width = 12

        # הוספת פילטרים
        if add_filters:
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

        # הקפאת השורה העליונה
        worksheet.freeze_panes = 'A2'

    def add_statistics_sheet(self, workbook, df: pd.DataFrame):
        """הוספת גיליון סטטיסטיקות"""
        ws = workbook.create_sheet(title="סטטיסטיקות")
        ws.sheet_view.rightToLeft = True

        stats_data = []

        # סטטיסטיקות בסיסיות
        stats_data.append(["סך הכל מילים", len(df)])
        stats_data.append(["מילים ייחודיות", df['מילה'].nunique()])

        # התפלגות סוגי הברות
        stats_data.append([""])
        stats_data.append(["התפלגות סוגי הברות", ""])
        for syllable_type, count in df['סוג הברה'].value_counts().items():
            stats_data.append([syllable_type, count])

        # מילים עם שווא
        stats_data.append([""])
        stats_data.append(["מילים עם שווא", df[df['יש שווא'] == 'כן'].shape[0]])
        stats_data.append(["מילים ללא שווא", df[df['יש שווא'] == 'לא'].shape[0]])

        # מילים עם דגש
        stats_data.append([""])
        stats_data.append(["מילים עם דגש", df[df['יש דגש'] == 'כן'].shape[0]])

        # מקרים מיוחדים
        stats_data.append([""])
        stats_data.append(["מקרים מיוחדים", ""])
        special_cases = df['מקרים מיוחדים'].str.split(', ', expand=True).stack()
        for case, count in special_cases.value_counts().items():
            if case:
                stats_data.append([case, count])

        # כתיבת הנתונים לגיליון
        for row_idx, row_data in enumerate(stats_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or (len(row_data) > 1 and col_idx == 1 and value and value != ""):
                    cell.font = Font(bold=True)
                cell.alignment = self.rtl_alignment

        # התאמת רוחב עמודות
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def add_pivot_sheet(self, workbook, df: pd.DataFrame):
        """הוספת גיליון עם טבלת Pivot"""
        # יצירת Pivot Table בסיסית
        pivot_data = df.pivot_table(
            index='סוג הברה',
            columns='יש שווא',
            values='מילה',
            aggfunc='count',
            fill_value=0
        )

        # יצירת גיליון חדש
        ws = workbook.create_sheet(title="טבלת צילוב")
        ws.sheet_view.rightToLeft = True

        # כתיבת הכותרות
        ws.cell(row=1, column=1, value="סוג הברה / יש שווא")
        for col_idx, col_name in enumerate(pivot_data.columns, 2):
            ws.cell(row=1, column=col_idx, value=col_name)

        # כתיבת הנתונים
        for row_idx, (index, row) in enumerate(pivot_data.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=index)
            for col_idx, value in enumerate(row, 2):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # עיצוב
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = self.center_alignment
                cell.border = self.border
                if cell.row == 1 or cell.column == 1:
                    cell.font = Font(bold=True)
                    if cell.row == 1:
                        cell.fill = self.header_fill
                        cell.font = self.header_font

    def add_info_sheet(self, workbook):
        """הוספת גיליון עם הסברים"""
        ws = workbook.create_sheet(title="הסברים")
        ws.sheet_view.rightToLeft = True

        info_text = [
            ["מדריך למשתמש", ""],
            ["", ""],
            ["שדות בגיליון התוצאות:", ""],
            ["מילה", "המילה המלאה עם ניקוד"],
            ["מילה ללא ניקוד", "המילה ללא סימני ניקוד"],
            ["סוג הברה", "פתוחה / סגורה / לא ידוע"],
            ["תבנית ניקוד", "ייצוג סכמטי של מבנה הניקוד"],
            ["יש שווא", "האם המילה מכילה שווא"],
            ["סוגי שווא", "נע / נח / שני שווא / וכו'"],
            ["סימני ניקוד", "רשימת כל סימני הניקוד במילה"],
            ["יש דגש", "האם המילה מכילה דגש"],
            ["הברה פתוחה", "האם יש במילה הברה פתוחה"],
            ["הברה סגורה", "האם יש במילה הברה סגורה"],
            ["מקרים מיוחדים", "קמץ קטן, פתח גנובה וכו'"],
            ["הקשר", "המשפט שבו המילה מופיעה"],
            ["מקור", "שם הטקסט/קובץ המקור"],
            ["קטגוריה", "סיווג הטקסט"],
            ["", ""],
            ["שימוש בפילטרים:", ""],
            ["1. לחץ על החץ בכותרת העמודה", ""],
            ["2. בחר את הערכים הרצויים", ""],
            ["3. ניתן לשלב מספר פילטרים", ""],
            ["", ""],
            ["טיפים:", ""],
            ["- השתמש ב-Ctrl+F לחיפוש מהיר", ""],
            ["- ניתן למיין כל עמודה בלחיצה על הכותרת", ""],
            ["- לייצוא חלק מהנתונים, סנן ואז העתק", ""]
        ]

        for row_idx, row_data in enumerate(info_text, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 1 and value and not row_data[1]:
                    cell.font = Font(bold=True, size=14)
                elif col_idx == 1:
                    cell.font = Font(bold=True)
                cell.alignment = self.rtl_alignment

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40

    def export_filtered_results(self, results: List[Dict], output_path: str,
                               columns: Optional[List[str]] = None):
        """ייצוא תוצאות עם עמודות נבחרות בלבד"""
        df = self.prepare_dataframe(results)

        if columns:
            # סינון רק העמודות המבוקשות
            available_columns = [col for col in columns if col in df.columns]
            df = df[available_columns]

        # ייצוא לאקסל
        self.export_to_excel(results, output_path)


# בדיקות בסיסיות
if __name__ == "__main__":
    # דוגמה לתוצאות חיפוש
    sample_results = [
        {
            'word': 'שָׁלוֹם',
            'word_plain': 'שלום',
            'syllable_type': 'פתוחה',
            'nikud_pattern': 'לתלתל',
            'has_shva': False,
            'shva_types': [],
            'nikud_marks': ['\u05B8', '\u05B9'],
            'has_dagesh': False,
            'has_open_syllable': True,
            'has_closed_syllable': False,
            'special_cases': [],
            'context': 'שָׁלוֹם לְכֻלָּם',
            'source_name': 'טקסט לדוגמה',
            'category_name': 'כללי',
            'position': 0
        },
        {
            'word': 'יְלָדִים',
            'word_plain': 'ילדים',
            'syllable_type': 'סגורה',
            'nikud_pattern': 'לשלתלתל',
            'has_shva': True,
            'shva_types': ['נע'],
            'nikud_marks': ['\u05B0', '\u05B8', '\u05B4'],
            'has_dagesh': False,
            'has_open_syllable': False,
            'has_closed_syllable': True,
            'special_cases': [],
            'context': 'הַיְלָדִים שִׂחֲקוּ בַּחָצֵר',
            'source_name': 'סיפור ילדים',
            'category_name': 'ספרות',
            'position': 5
        }
    ]

    exporter = ExcelExporter()
    output_file = "תוצאות_חיפוש_ניקוד.xlsx"

    print(f"מייצא תוצאות לקובץ {output_file}...")
    exporter.export_to_excel(sample_results, output_file, add_filters=True, add_pivot=True)
    print("הייצוא הושלם בהצלחה!")